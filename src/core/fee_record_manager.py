"""
Fee Record Manager - Core Excel operations for loading table data to fee record
Handles month matching, merged cell detection, and data insertion with yellow highlighting
Dynamic month detection and creation without hardcoding
File: src/core/fee_record_manager.py
"""
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from typing import List, Dict, Tuple, Optional, Any, Set
from datetime import datetime
import os
import shutil


class FeeRecordManager:
    """Manages loading preview table data into Fee Record Excel file with dynamic month detection"""
    
    # Month name mapping from preview table to fee record headers
    MONTH_MAPPING = {
        "Jan": "JANUARY", "Feb": "FEBRUARY", "Mar": "MARCH",
        "Apr": "APRIL", "May": "MAY", "Jun": "JUNE", 
        "Jul": "JULY", "Aug": "AUGUST", "Sep": "SEPTEMBER",
        "Oct": "OCTOBER", "Nov": "NOVEMBER", "Dec": "DECEMBER"
    }
    
    # Chronological month order for proper insertion
    MONTH_ORDER = {
        "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
        "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
        "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12
    }
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.fee_record_path = ""
        self.column_mapping = {}  # {month: {"merged_range": (start, end), "date_col": col, "amount_col": col}}
        self.parent_column = 1    # Assume first column has parent names
        
        # Cell styling for highlighting updated cells
        self.highlight_fill = PatternFill(
            start_color="FFFF00",  # Yellow background
            end_color="FFFF00",
            fill_type="solid"
        )
        
        # Light blue fill for new parent names
        self.new_parent_fill = PatternFill(
            start_color="E6F3FF",  # Light blue background
            end_color="E6F3FF",
            fill_type="solid"
        )
        
        # Track updated cells for statistics
        self.updated_cells = []
        
    def load_table_data_to_fee_record(self, table_data: List[List[str]], 
                                     fee_record_file_path: str) -> Dict[str, Any]:
        """
        Main method to load preview table data into fee record file
        Enhanced with dynamic month detection and cell highlighting
        
        Args:
            table_data: Current table data from get_all_data()
            fee_record_file_path: Path to fee record Excel file
            
        Returns:
            Dict with success status, statistics, and highlighting info
        """
        try:
            # Validate inputs
            if not table_data:
                return {"success": False, "error": "No data to load"}
                
            if not os.path.exists(fee_record_file_path):
                return {"success": False, "error": f"Fee record file not found: {fee_record_file_path}"}
            
            # Check if file is locked/in use
            if self._is_file_locked(fee_record_file_path):
                return {
                    "success": False, 
                    "error": f"File is currently in use or locked.\n\n"
                            f"Please:\n"
                            f"1. Close Excel if the file is open\n"
                            f"2. Make sure no other program is using the file\n"
                            f"3. Check that you have write permissions\n\n"
                            f"File: {fee_record_file_path}"
                }
            
            # Create backup
            backup_path = self._create_backup(fee_record_file_path)
            
            # Load Excel file with error handling
            try:
                self.fee_record_path = fee_record_file_path
                self.workbook = openpyxl.load_workbook(fee_record_file_path)
                self.worksheet = self.workbook.active
            except PermissionError as e:
                return {
                    "success": False,
                    "error": f"Permission denied accessing file.\n\n"
                            f"Please:\n"
                            f"1. Close Excel if the file is open\n"
                            f"2. Run the application as administrator if needed\n"
                            f"3. Check file permissions\n\n"
                            f"File: {fee_record_file_path}\n"
                            f"Error: {str(e)}"
                }
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Failed to open Excel file.\n\n"
                            f"Error: {str(e)}\n"
                            f"File: {fee_record_file_path}"
                }
            
            # Analyze existing structure dynamically
            self._analyze_fee_record_structure()
            
            # Process table data (with highlighting)
            stats = self._process_table_data(table_data)
            
            # Get highlighting summary
            highlighting_summary = self.get_highlighting_summary()
            
            # Save changes with error handling
            try:
                self.workbook.save(fee_record_file_path)
            except PermissionError as e:
                return {
                    "success": False,
                    "error": f"Permission denied saving file.\n\n"
                            f"The file may be open in Excel or write-protected.\n"
                            f"Please close Excel and try again.\n\n"
                            f"File: {fee_record_file_path}\n"
                            f"Error: {str(e)}"
                }
            
            return {
                "success": True,
                "backup_path": backup_path,
                "stats": stats,
                "highlighting": highlighting_summary
            }
            
        except Exception as e:
            return {"success": False, "error": f"Unexpected error: {str(e)}"}
        finally:
            if self.workbook:
                try:
                    self.workbook.close()
                except:
                    pass

    def _create_backup(self, file_path: str) -> str:
        """Create backup of original fee record file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{file_path}.backup_{timestamp}"
        shutil.copy2(file_path, backup_path)
        return backup_path
    
    def _is_file_locked(self, file_path: str) -> bool:
        """
        Check if file is locked or in use by another process
        
        Args:
            file_path: Path to file to check
            
        Returns:
            bool: True if file is locked, False otherwise
        """
        try:
            # Try to open file in append mode (less intrusive than write mode)
            with open(file_path, 'a'):
                pass
            return False
        except IOError:
            return True
        except Exception:
            return True
    
    def _analyze_fee_record_structure(self):
        """
        Dynamically analyze fee record file structure and build column mapping
        Detects merged month headers and maps to date/amount column pairs
        """
        if not self.worksheet:
            raise Exception("Worksheet not loaded")
        
        self.column_mapping = {}
        
        # Get all merged cell ranges in header row (row 1)
        merged_ranges = list(self.worksheet.merged_cells.ranges)
        
        # Check each merged range to see if it contains a month name
        for merged_range in merged_ranges:
            # Only process ranges that include row 1 (header row)
            if merged_range.min_row <= 1 <= merged_range.max_row:
                # Get the value from the top-left cell of the merged range
                header_cell = self.worksheet.cell(row=1, column=merged_range.min_col)
                header_value = header_cell.value
                
                if header_value:
                    header_text = str(header_value).strip().upper()
                    
                    # Check if this is a month name
                    if header_text in self.MONTH_ORDER:
                        # Calculate expected columns (should span exactly 2 columns)
                        start_col = merged_range.min_col
                        end_col = merged_range.max_col
                        
                        # Verify it spans exactly 2 columns
                        if end_col - start_col + 1 == 2:
                            self.column_mapping[header_text] = {
                                "merged_range": (start_col, end_col),
                                "date_col": start_col,      # First column is date
                                "amount_col": start_col + 1  # Second column is amount
                            }
        
        # Also check for non-merged month headers (fallback)
        self._detect_non_merged_months()
        
        detected_count = len(self.column_mapping)
        detected_months = list(self.column_mapping.keys())
        print(f"Detected {detected_count} month columns: {detected_months}")
    
    def _detect_non_merged_months(self):
        """
        Detect non-merged month headers as fallback
        Some fee records might not use merged cells
        """
        # Scan header row for month names that aren't already mapped
        for col in range(1, self.worksheet.max_column + 1):
            header_cell = self.worksheet.cell(row=1, column=col)
            if header_cell.value:
                header_text = str(header_cell.value).strip().upper()
                
                # Check if this is a month name we haven't mapped yet
                if header_text in self.MONTH_ORDER and header_text not in self.column_mapping:
                    # Assume next column is the amount column
                    if col + 1 <= self.worksheet.max_column:
                        self.column_mapping[header_text] = {
                            "merged_range": (col, col + 1),
                            "date_col": col,
                            "amount_col": col + 1
                        }
    
    def _process_table_data(self, table_data: List[List[str]]) -> Dict[str, int]:
        """Process all table data and insert into fee record with highlighting"""
        stats = {
            "processed_rows": 0,
            "updated_entries": 0,
            "new_parents": 0,
            "new_months_created": 0,
            "errors": 0,
            "highlighted_cells": 0
        }
        
        # Clear updated cells tracking
        self.updated_cells = []
        
        # Group data by required months to create missing ones first
        required_months = set()
        for row in table_data:
            if len(row) >= 5 and row[4]:  # Month Paying For column
                month_3letter = row[4].strip()
                if month_3letter in self.MONTH_MAPPING:
                    required_months.add(self.MONTH_MAPPING[month_3letter])
        
        # Create missing month columns and re-scan structure after each creation
        months_created = 0
        for month in required_months:
            if month not in self.column_mapping:
                print(f"Creating missing month: {month}")
                self._create_month_column(month)
                months_created += 1
                
                # Re-analyze structure after creating new month
                print(f"Re-scanning structure after creating {month}")
                self._analyze_fee_record_structure()
                current_months = list(self.column_mapping.keys())
                print(f"Updated column mapping: {current_months}")
                
                stats["new_months_created"] += 1
        
        # Verify all required months are now available
        missing_months = []
        for month in required_months:
            if month not in self.column_mapping:
                missing_months.append(month)
        
        if missing_months:
            print(f"WARNING: Still missing months after creation: {missing_months}")
        
        # Process each data row
        for row_data in table_data:
            try:
                if self._process_single_row(row_data):
                    stats["updated_entries"] += 1
                stats["processed_rows"] += 1
            except Exception as e:
                print(f"Error processing row {row_data}: {e}")
                stats["errors"] += 1
        
        # Update highlighted cells count
        stats["highlighted_cells"] = len(self.updated_cells)
        
        return stats
    
    def _create_month_column(self, month_name: str):
        """
        Create new month column with merged header spanning 2 columns (date and amount)
        Inserts in REVERSE chronological order (latest months on the LEFT)
        
        Args:
            month_name: Full month name (e.g., "SEPTEMBER", "OCTOBER")
        """
        # Find correct insertion point based on reverse chronological order
        insertion_col = self._find_month_insertion_point(month_name)
        
        # Insert 2 new columns (date and amount) at the insertion point
        self.worksheet.insert_cols(insertion_col, 2)
        
        # Create merged header cell spanning both columns
        month_header_cell = self.worksheet.cell(row=1, column=insertion_col, value=month_name)
        month_header_cell.font = Font(bold=True)
        
        # Merge and center the header across both columns  
        self.worksheet.merge_cells(
            start_row=1, start_column=insertion_col,
            end_row=1, end_column=insertion_col + 1
        )
        
        # Apply center alignment to the merged cell
        month_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Update column mapping for the new month
        self.column_mapping[month_name] = {
            "merged_range": (insertion_col, insertion_col + 1),
            "date_col": insertion_col,      # First column is date
            "amount_col": insertion_col + 1  # Second column is amount
        }
        
        # Update existing column mappings (all shift right by 2)
        self._shift_column_mappings_after_insertion(insertion_col, 2)
        
        start_col = insertion_col
        end_col = insertion_col + 1
        print(f"Created new month column {month_name} at columns {start_col}-{end_col} (merged and centered)")
        print("All existing months shifted RIGHT by 2 columns")
    
    def _find_month_insertion_point(self, month_name: str) -> int:
        """
        Find where to insert new month column based on REVERSE chronological order
        New months are inserted to the LEFT of existing months (latest months first)
        
        Args:
            month_name: Month name to insert
            
        Returns:
            int: Column index where to insert the new month
        """
        target_order = self.MONTH_ORDER.get(month_name, 0)
        
        # If no existing months, insert after parent column
        if not self.column_mapping:
            return self.parent_column + 1
        
        # Find the leftmost (earliest in year) existing month
        # Since months are in reverse order, we want to insert BEFORE months that come later in the year
        insertion_point = None
        leftmost_col = float('inf')
        
        for existing_month, mapping in self.column_mapping.items():
            existing_order = self.MONTH_ORDER.get(existing_month, 0)
            
            # If target month should come AFTER this existing month in the year (but LEFT in the spreadsheet)
            if target_order > existing_order:
                # Insert before this month (to the left)
                if mapping["date_col"] < leftmost_col:
                    leftmost_col = mapping["date_col"]
                    insertion_point = mapping["date_col"]
        
        # If no insertion point found (target month is earliest in year), 
        # insert at the leftmost position (after parent column)
        if insertion_point is None:
            # Find the leftmost existing month column
            leftmost_col = float('inf')
            for mapping in self.column_mapping.values():
                leftmost_col = min(leftmost_col, mapping["date_col"])
            
            # Insert before the leftmost existing month
            insertion_point = leftmost_col if leftmost_col != float('inf') else self.parent_column + 1
        
        return insertion_point
    
    def _shift_column_mappings_after_insertion(self, insertion_col: int, cols_inserted: int):
        """
        Update existing column mappings after inserting new columns
        Since we insert to the LEFT, ALL existing columns shift RIGHT
        
        Args:
            insertion_col: Column where insertion occurred
            cols_inserted: Number of columns inserted
        """
        for month, mapping in self.column_mapping.items():
            # All existing columns at or after the insertion point shift right
            if mapping["date_col"] >= insertion_col:
                # Update merged range
                old_start, old_end = mapping["merged_range"]
                mapping["merged_range"] = (old_start + cols_inserted, old_end + cols_inserted)
                
                # Update date and amount columns
                mapping["date_col"] += cols_inserted
                mapping["amount_col"] += cols_inserted
    
    def _process_single_row(self, row_data: List[str]) -> bool:
        """Process a single row from table data with yellow cell highlighting"""
        if len(row_data) < 6:
            return False
        
        # Extract data
        parent_name = row_data[2].strip() if row_data[2] else ""  # Matched Parent
        month_3letter = row_data[4].strip() if row_data[4] else ""  # Month Paying For
        transaction_date = row_data[1].strip() if row_data[1] else ""  # Transaction Date
        amount = row_data[5].strip() if row_data[5] else ""  # Amount
        
        if not all([parent_name, month_3letter]):
            return False
        
        # Convert month format
        month_full = self.MONTH_MAPPING.get(month_3letter)
        if not month_full:
            print(f"ERROR: Cannot map month {month_3letter} to full month name")
            return False
        
        # Verify the month exists in current mapping
        if month_full not in self.column_mapping:
            print(f"ERROR: Month {month_full} not found in column mapping!")
            available = list(self.column_mapping.keys())
            print(f"Available months: {available}")
            return False
        
        parent_info = parent_name
        month_info = month_full
        date_info = transaction_date
        amount_info = amount
        print(f"Processing: Parent={parent_info}, Month={month_info}, Date={date_info}, Amount={amount_info}")
        
        # Find or create parent row
        parent_row = self._find_or_create_parent_row(parent_name)
        
        # Find next available row in this month's columns for this parent
        target_row = self._find_next_available_row_in_month(month_full, parent_row)
        
        # Get month column mapping (should be current/updated mapping)
        month_cols = self.column_mapping[month_full]
        date_col_info = month_cols["date_col"]
        amount_col_info = month_cols["amount_col"]
        print(f"Using month {month_full} columns: date_col={date_col_info}, amount_col={amount_col_info}")
        
        cells_updated = 0
        
        # Insert date with yellow highlighting
        if transaction_date and "date_col" in month_cols:
            date_cell = self.worksheet.cell(row=target_row, column=month_cols["date_col"], 
                                          value=transaction_date)
            date_cell.fill = self.highlight_fill
            
            self.updated_cells.append({
                'row': target_row, 'col': month_cols["date_col"],
                'type': 'date', 'value': transaction_date,
                'parent': parent_name, 'month': month_full
            })
            cells_updated += 1
            row_info = target_row
            col_info = month_cols["date_col"]
            print(f"  -> Inserted date {transaction_date} at row {row_info}, col {col_info}")
        
        # Insert amount with yellow highlighting
        if amount and "amount_col" in month_cols:
            try:
                # Try to convert to float for proper formatting
                amount_float = float(amount.replace(',', '').replace('$', '').replace('RM', ''))
                amount_cell = self.worksheet.cell(row=target_row, column=month_cols["amount_col"], 
                                                value=amount_float)
            except ValueError:
                # If conversion fails, insert as text
                amount_cell = self.worksheet.cell(row=target_row, column=month_cols["amount_col"], 
                                                value=amount)
            
            amount_cell.fill = self.highlight_fill
            
            self.updated_cells.append({
                'row': target_row, 'col': month_cols["amount_col"],
                'type': 'amount', 'value': amount,
                'parent': parent_name, 'month': month_full
            })
            cells_updated += 1
            row_info = target_row
            col_info = month_cols["amount_col"]
            print(f"  -> Inserted amount {amount} at row {row_info}, col {col_info}")
        
        parent_info = parent_name
        month_info = month_full
        cells_info = cells_updated
        print(f"  -> Updated {cells_info} cells for {parent_info} in {month_info}")
        return cells_updated > 0
    
    def _find_or_create_parent_row(self, parent_name: str) -> int:
        """Find existing parent row or create new one"""
        # Search existing rows for this parent
        for row in range(2, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=self.parent_column).value
            if cell_value and str(cell_value).strip().upper() == parent_name.upper():
                return row
        
        # Create new row at the end
        new_row = self.worksheet.max_row + 1
        parent_cell = self.worksheet.cell(row=new_row, column=self.parent_column, value=parent_name)
        parent_cell.fill = self.new_parent_fill  # Light blue for new parents
        
        return new_row
    
    def _find_next_available_row_in_month(self, month_name: str, preferred_row: int) -> int:
        """
        Find the next available row in the month's columns
        Starts from the parent's row and looks for empty cells
        
        Args:
            month_name: Month to check
            preferred_row: Preferred row (parent's row)
            
        Returns:
            int: Row number to use for data insertion
        """
        if month_name not in self.column_mapping:
            return preferred_row
        
        month_cols = self.column_mapping[month_name]
        date_col = month_cols["date_col"]
        amount_col = month_cols["amount_col"]
        
        # Start checking from the preferred row downward
        for row in range(preferred_row, self.worksheet.max_row + 2):
            # Check if both date and amount cells are empty
            date_cell = self.worksheet.cell(row=row, column=date_col).value
            amount_cell = self.worksheet.cell(row=row, column=amount_col).value
            
            if not date_cell and not amount_cell:
                return row
        
        # If no empty row found, use the next row after max_row
        return self.worksheet.max_row + 1
    
    def get_highlighting_summary(self) -> Dict[str, Any]:
        """Get summary of highlighted cells for user feedback"""
        if not self.updated_cells:
            return {"total_highlighted": 0, "by_type": {}, "by_parent": {}, "by_month": {}}
        
        # Group by type (date vs amount)
        by_type = {"date": 0, "amount": 0}
        by_parent = {}
        by_month = {}
        
        for cell_info in self.updated_cells:
            # Count by type
            cell_type = cell_info['type']
            by_type[cell_type] = by_type.get(cell_type, 0) + 1
            
            # Count by parent
            parent = cell_info['parent']
            by_parent[parent] = by_parent.get(parent, 0) + 1
            
            # Count by month
            month = cell_info['month']
            by_month[month] = by_month.get(month, 0) + 1
        
        return {
            "total_highlighted": len(self.updated_cells),
            "by_type": by_type,
            "by_parent": by_parent,
            "by_month": by_month,
            "details": self.updated_cells
        }
    
    def preview_changes(self, table_data: List[List[str]], 
                       fee_record_file_path: str) -> Dict[str, Any]:
        """Preview what changes will be made without actually modifying the file"""
        try:
            # Load file in WRITE mode to access merged_cells (but don't save)
            temp_workbook = openpyxl.load_workbook(fee_record_file_path)
            temp_worksheet = temp_workbook.active
            
            # Analyze what would happen
            preview_info = {
                "total_rows": len(table_data),
                "new_months": [],
                "affected_parents": set(),
                "data_updates": []
            }
            
            # Detect existing months using the same logic as main analysis
            existing_months = set()
            
            # Check merged cells if available
            if hasattr(temp_worksheet, 'merged_cells'):
                merged_ranges = list(temp_worksheet.merged_cells.ranges)
                
                for merged_range in merged_ranges:
                    if merged_range.min_row <= 1 <= merged_range.max_row:
                        header_cell = temp_worksheet.cell(row=1, column=merged_range.min_col)
                        if header_cell.value:
                            header_text = str(header_cell.value).strip().upper()
                            if header_text in self.MONTH_ORDER:
                                existing_months.add(header_text)
            
            # Fallback: scan header row for month names (non-merged)
            for col in range(1, temp_worksheet.max_column + 1):
                header_cell = temp_worksheet.cell(row=1, column=col)
                if header_cell.value:
                    header_text = str(header_cell.value).strip().upper()
                    if header_text in self.MONTH_ORDER:
                        existing_months.add(header_text)
            
            # Find new months required
            for row in table_data:
                if len(row) >= 5 and row[4]:
                    month_3letter = row[4].strip()
                    if month_3letter in self.MONTH_MAPPING:
                        month_full = self.MONTH_MAPPING[month_3letter]
                        if month_full not in existing_months:
                            preview_info["new_months"].append(month_full)
                        
                        if len(row) >= 3 and row[2]:
                            preview_info["affected_parents"].add(row[2].strip())
            
            preview_info["new_months"] = list(set(preview_info["new_months"]))
            preview_info["affected_parents"] = list(preview_info["affected_parents"])
            
            # Close without saving
            temp_workbook.close()
            return preview_info
            
        except Exception as e:
            return {"error": str(e)}
    
    def validate_table_data(self, table_data: List[List[str]]) -> List[str]:
        """Validate table data before loading"""
        errors = []
        
        if not table_data:
            errors.append("No data to validate")
            return errors
        
        for i, row in enumerate(table_data):
            row_num = i + 1
            
            # Check row length
            if len(row) < 6:
                errors.append(f"Row {row_num}: Missing columns (expected 6)")
                continue
            
            # Check required fields
            if not row[2] or not row[2].strip():  # Matched Parent
                errors.append(f"Row {row_num}: Missing parent name")
            
            if not row[4] or not row[4].strip():  # Month Paying For
                errors.append(f"Row {row_num}: Missing month")
            elif row[4].strip() not in self.MONTH_MAPPING:
                month_val = row[4]
                errors.append(f"Row {row_num}: Invalid month {month_val}")
            
            # Validate amount if present
            if row[5] and row[5].strip():
                try:
                    float(row[5].replace(',', '').replace('$', '').replace('RM', ''))
                except ValueError:
                    amount_val = row[5]
                    errors.append(f"Row {row_num}: Invalid amount format {amount_val}")
        
        return errors