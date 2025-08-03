# File: src/gui/outstanding_payments_tab/payment_export.py
"""
Payment Export Manager - Handles exporting outstanding payment data to various formats
Provides Excel, CSV, and PDF export capabilities with customizable formatting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from datetime import datetime
import os
import csv
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from PyQt5.QtCore import QObject


class PaymentExporter(QObject):
    """
    Handles exporting outstanding payment data to various formats
    Supports Excel, CSV, and formatted reports
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Styling for Excel exports
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.unpaid_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        self.paid_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
        
        # Border styling
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_outstanding_payments_excel(self, analysis_data: Dict[str, Any], 
                                        parent_widget=None) -> bool:
        """
        Export outstanding payments to Excel with formatting
        
        Args:
            analysis_data: Payment analysis results from PaymentAnalyzer
            parent_widget: Parent widget for file dialog
            
        Returns:
            bool: True if export successful, False otherwise
        """
        try:
            # Generate default filename
            month_name = analysis_data.get('month_display', 'Outstanding')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Outstanding_Payments_{month_name}_{timestamp}.xlsx"
            
            # Show save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Export Outstanding Payments",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return False  # User cancelled
            
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"Outstanding {month_name}"
            
            # Write header information
            self._write_excel_header(worksheet, analysis_data)
            
            # Write outstanding payments table
            table_start_row = self._write_outstanding_table(worksheet, analysis_data)
            
            # Write summary information
            self._write_summary_section(worksheet, analysis_data, table_start_row + 3)
            
            # Apply formatting
            self._format_excel_worksheet(worksheet, analysis_data)
            
            # Save file
            workbook.save(file_path)
            workbook.close()
            
            QMessageBox.information(
                parent_widget, 
                "Export Successful", 
                f"Outstanding payments exported to:\n{file_path}"
            )
            
            return True
            
        except Exception as e:
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                f"Failed to export outstanding payments:\n{str(e)}"
            )
            return False
    
    def export_outstanding_payments_csv(self, analysis_data: Dict[str, Any],
                                      parent_widget=None) -> bool:
        """
        Export outstanding payments to CSV format
        
        Args:
            analysis_data: Payment analysis results from PaymentAnalyzer
            parent_widget: Parent widget for file dialog
            
        Returns:
            bool: True if export successful, False otherwise
        """
        try:
            # Generate default filename
            month_name = analysis_data.get('month_display', 'Outstanding')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Outstanding_Payments_{month_name}_{timestamp}.csv"
            
            # Show save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Export Outstanding Payments (CSV)",
                default_filename,
                "CSV Files (*.csv);;All Files (*)"
            )
            
            if not file_path:
                return False  # User cancelled
            
            # Prepare CSV data
            csv_data = self._prepare_csv_data(analysis_data)
            
            # Write CSV file
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header row
                writer.writerow(csv_data['headers'])
                
                # Write data rows
                writer.writerows(csv_data['rows'])
            
            QMessageBox.information(
                parent_widget,
                "Export Successful",
                f"Outstanding payments exported to:\n{file_path}"
            )
            
            return True
            
        except Exception as e:
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                f"Failed to export CSV:\n{str(e)}"
            )
            return False
    
    def export_summary_report(self, analysis_data: Dict[str, Any],
                            parent_widget=None) -> bool:
        """
        Export a formatted summary report with both paid and unpaid parents
        
        Args:
            analysis_data: Payment analysis results from PaymentAnalyzer
            parent_widget: Parent widget for file dialog
            
        Returns:
            bool: True if export successful, False otherwise
        """
        try:
            # Generate default filename
            month_name = analysis_data.get('month_display', 'Payment')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Payment_Summary_{month_name}_{timestamp}.xlsx"
            
            # Show save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Export Payment Summary Report",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return False  # User cancelled
            
            # Create workbook with multiple sheets
            workbook = openpyxl.Workbook()
            
            # Summary sheet
            summary_sheet = workbook.active
            summary_sheet.title = "Summary"
            self._create_summary_sheet(summary_sheet, analysis_data)
            
            # Outstanding payments sheet
            outstanding_sheet = workbook.create_sheet("Outstanding Payments")
            self._create_outstanding_sheet(outstanding_sheet, analysis_data)
            
            # Paid payments sheet
            paid_sheet = workbook.create_sheet("Paid Payments")
            self._create_paid_sheet(paid_sheet, analysis_data)
            
            # Save file
            workbook.save(file_path)
            workbook.close()
            
            QMessageBox.information(
                parent_widget,
                "Export Successful",
                f"Payment summary report exported to:\n{file_path}"
            )
            
            return True
            
        except Exception as e:
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                f"Failed to export summary report:\n{str(e)}"
            )
            return False
    
    def _write_excel_header(self, worksheet, analysis_data: Dict[str, Any]):
        """Write header information to Excel worksheet"""
        month_name = analysis_data.get('month_display', 'Unknown')
        
        # Title
        worksheet['A1'] = f"Outstanding Payments Report - {month_name}"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Report details
        worksheet['A3'] = f"Month: {month_name}"
        worksheet['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet['A5'] = f"Total Parents: {analysis_data.get('total_parents', 0)}"
        worksheet['A6'] = f"Outstanding: {analysis_data.get('unpaid_count', 0)}"
        
        return 7  # Next available row
    
    def _write_outstanding_table(self, worksheet, analysis_data: Dict[str, Any]) -> int:
        """Write outstanding payments table to worksheet"""
        unpaid_parents = analysis_data.get('unpaid_parents', [])
        start_row = 8
        
        # Headers
        headers = ['No.', 'Parent Name', 'Student Name', 'Date Value', 'Amount Value', 'Status']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for idx, parent_data in enumerate(unpaid_parents, 1):
            row = start_row + idx
            
            # Row data
            row_data = [
                idx,
                parent_data.get('parent_name', ''),
                parent_data.get('student_name', ''),
                parent_data.get('date_value', ''),
                parent_data.get('formatted_amount', ''),
                'Outstanding'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.fill = self.unpaid_fill
                cell.border = self.thin_border
                
                # Center align certain columns
                if col in [1, 6]:  # Number and Status columns
                    cell.alignment = Alignment(horizontal='center')
        
        return start_row + len(unpaid_parents)
    
    def _write_summary_section(self, worksheet, analysis_data: Dict[str, Any], start_row: int):
        """Write summary statistics section"""
        total_parents = analysis_data.get('total_parents', 0)
        paid_count = analysis_data.get('paid_count', 0)
        unpaid_count = analysis_data.get('unpaid_count', 0)
        
        # Summary statistics
        worksheet.cell(row=start_row, column=1, value="Summary Statistics").font = Font(bold=True)
        worksheet.cell(row=start_row + 1, column=1, value=f"Total Parents: {total_parents}")
        worksheet.cell(row=start_row + 2, column=1, value=f"Paid: {paid_count}")
        worksheet.cell(row=start_row + 3, column=1, value=f"Outstanding: {unpaid_count}")
        
        if total_parents > 0:
            payment_rate = (paid_count / total_parents) * 100
            worksheet.cell(row=start_row + 4, column=1, value=f"Payment Rate: {payment_rate:.1f}%")
    
    def _format_excel_worksheet(self, worksheet, analysis_data: Dict[str, Any]):
        """Apply formatting to Excel worksheet"""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _prepare_csv_data(self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for CSV export"""
        unpaid_parents = analysis_data.get('unpaid_parents', [])
        month_name = analysis_data.get('month_display', 'Unknown')
        
        # Headers
        headers = [
            'Parent Name',
            'Student Name', 
            'Month',
            'Date Value',
            'Amount Value',
            'Status',
            'Export Date'
        ]
        
        # Data rows
        rows = []
        export_date = datetime.now().strftime('%Y-%m-%d')
        
        for parent_data in unpaid_parents:
            row = [
                parent_data.get('parent_name', ''),
                parent_data.get('student_name', ''),
                month_name,
                parent_data.get('date_value', ''),
                parent_data.get('formatted_amount', ''),
                'Outstanding',
                export_date
            ]
            rows.append(row)
        
        return {'headers': headers, 'rows': rows}
    
    def _create_summary_sheet(self, worksheet, analysis_data: Dict[str, Any]):
        """Create summary overview sheet"""
        month_name = analysis_data.get('month_display', 'Unknown')
        
        # Title
        worksheet['A1'] = f"Payment Summary - {month_name}"
        worksheet['A1'].font = Font(size=18, bold=True)
        
        # Statistics
        stats_start = 3
        stats_data = [
            ('Total Parents', analysis_data.get('total_parents', 0)),
            ('Paid', analysis_data.get('paid_count', 0)),
            ('Outstanding', analysis_data.get('unpaid_count', 0)),
            ('Payment Rate', f"{(analysis_data.get('paid_count', 0) / max(analysis_data.get('total_parents', 1), 1) * 100):.1f}%")
        ]
        
        for idx, (label, value) in enumerate(stats_data):
            worksheet.cell(row=stats_start + idx, column=1, value=label).font = Font(bold=True)
            worksheet.cell(row=stats_start + idx, column=2, value=value)
        
        # Report info
        info_start = stats_start + len(stats_data) + 2
        worksheet.cell(row=info_start, column=1, value="Report Generated:").font = Font(bold=True)
        worksheet.cell(row=info_start, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    def _create_outstanding_sheet(self, worksheet, analysis_data: Dict[str, Any]):
        """Create outstanding payments detailed sheet"""
        unpaid_parents = analysis_data.get('unpaid_parents', [])
        
        # Headers
        headers = ['Parent Name', 'Student Name', 'Date Value', 'Amount Value', 'Row in Fee Record']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        # Data
        for idx, parent_data in enumerate(unpaid_parents, 2):
            row_data = [
                parent_data.get('parent_name', ''),
                parent_data.get('student_name', ''),
                parent_data.get('date_value', ''),
                parent_data.get('formatted_amount', ''),
                parent_data.get('row', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=idx, column=col, value=value)
                cell.fill = self.unpaid_fill
    
    def _create_paid_sheet(self, worksheet, analysis_data: Dict[str, Any]):
        """Create paid payments detailed sheet"""
        paid_parents = analysis_data.get('paid_parents', [])
        
        # Headers  
        headers = ['Parent Name', 'Student Name', 'Date Value', 'Amount Value', 'Row in Fee Record']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        # Data
        for idx, parent_data in enumerate(paid_parents, 2):
            row_data = [
                parent_data.get('parent_name', ''),
                parent_data.get('student_name', ''),
                parent_data.get('date_value', ''),
                parent_data.get('formatted_amount', ''),
                parent_data.get('row', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=idx, column=col, value=value)
                cell.fill = self.paid_fill
    
    def get_export_options(self) -> List[Dict[str, str]]:
        """
        Get available export options for UI
        
        Returns:
            List of export option dictionaries
        """
        return [
            {
                'name': 'Outstanding Only (Excel)',
                'description': 'Export only unpaid parents to Excel with formatting',
                'method': 'excel_outstanding'
            },
            {
                'name': 'Outstanding Only (CSV)',
                'description': 'Export only unpaid parents to CSV format',
                'method': 'csv_outstanding'
            },
            {
                'name': 'Complete Summary Report',
                'description': 'Export complete report with paid and unpaid parents',
                'method': 'summary_report'
            }
        ]