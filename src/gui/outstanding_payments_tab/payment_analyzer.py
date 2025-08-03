# File: src/gui/outstanding_payments_tab/payment_analyzer.py
"""
Payment Analyzer - Core logic for analyzing fee records and identifying outstanding payments
Reuses the month detection logic from FeeRecordManager for consistency
FIXED: Handle None values properly to avoid comparison errors
"""

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from typing import List, Dict, Tuple, Optional, Any, Set
from datetime import datetime, date
import os
import re


class PaymentAnalyzer:
    """
    Analyzes fee record Excel files to identify parents with outstanding payments
    Uses the same month detection logic as FeeRecordManager for consistency
    """
    
    # Month name mapping from FeeRecordManager
    MONTH_MAPPING = {
        "Jan": "JANUARY", "Feb": "FEBRUARY", "Mar": "MARCH",
        "Apr": "APRIL", "May": "MAY", "Jun": "JUNE", 
        "Jul": "JULY", "Aug": "AUGUST", "Sep": "SEPTEMBER",
        "Oct": "OCTOBER", "Nov": "NOVEMBER", "Dec": "DECEMBER"
    }
    
    # Reverse mapping for display
    MONTH_DISPLAY = {v: k for k, v in MONTH_MAPPING.items()}
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.fee_record_path = ""
        self.column_mapping = {}  # Same structure as FeeRecordManager
        self.parent_column = 1  # Column A contains parent names
        
    def load_fee_record(self, file_path: str) -> bool:
        """
        Load fee record Excel file and analyze structure
        Uses same logic as FeeRecordManager._analyze_fee_record_structure()
        
        Args:
            file_path: Path to Excel fee record file
            
        Returns:
            bool: True if loaded successfully, False otherwise
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Fee record file not found: {file_path}")
                
            self.fee_record_path = file_path
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.worksheet = self.workbook.active
            
            # Use the same month detection logic as FeeRecordManager
            self._analyze_fee_record_structure()
            
            return True
            
        except Exception as e:
            print(f"Error loading fee record: {e}")
            return False
    
    def _analyze_fee_record_structure(self):
        """
        Dynamically analyze fee record file structure and build column mapping
        COPIED from FeeRecordManager._analyze_fee_record_structure() for consistency
        """
        if not self.worksheet:
            raise Exception("Worksheet not loaded")
        
        self.column_mapping = {}
        merged_ranges = list(self.worksheet.merged_cells.ranges)
        
        # Process merged month headers first
        for merged_range in merged_ranges:
            if merged_range.min_row <= 1 <= merged_range.max_row:
                header_cell = self.worksheet.cell(row=1, column=merged_range.min_col)
                header_value = header_cell.value
                
                if header_value:
                    header_text = str(header_value).strip().upper()
                    
                    # Check if this is a month header (using same logic as FeeRecordManager)
                    if header_text in self.MONTH_MAPPING.values():
                        start_col = merged_range.min_col
                        end_col = merged_range.max_col
                        
                        # Assume merged month spans 2 columns: date and amount
                        if end_col - start_col + 1 == 2:
                            self.column_mapping[header_text] = {
                                "merged_range": (start_col, end_col),
                                "date_col": start_col,
                                "amount_col": start_col + 1
                            }
        
        # Detect non-merged month headers as fallback
        self._detect_non_merged_months()
    
    def _detect_non_merged_months(self):
        """
        Detect non-merged month headers as fallback
        COPIED from FeeRecordManager._detect_non_merged_months() for consistency
        """
        for col in range(1, self.worksheet.max_column + 1):
            header_cell = self.worksheet.cell(row=1, column=col)
            if header_cell.value:
                header_text = str(header_cell.value).strip().upper()
                
                # Check if this is a month header and not already mapped
                if (header_text in self.MONTH_MAPPING.values() and 
                    header_text not in self.column_mapping):
                    
                    if col + 1 <= self.worksheet.max_column:
                        self.column_mapping[header_text] = {
                            "merged_range": (col, col + 1),
                            "date_col": col,
                            "amount_col": col + 1
                        }
    
    def get_available_months(self) -> List[str]:
        """
        Get list of available months in the fee record
        
        Returns:
            List of month names found in the fee record (e.g., ["DECEMBER", "JULY"])
        """
        return list(self.column_mapping.keys())
    
    def get_available_months_display(self) -> List[str]:
        """
        Get list of available months in user-friendly format
        
        Returns:
            List of month names in short format (e.g., ["Dec", "Jul"])
        """
        return [self.MONTH_DISPLAY.get(month, month) for month in self.get_available_months()]
    
    def get_all_parents(self) -> List[Dict[str, Any]]:
        """
        Get list of all parents from the fee record
        
        Returns:
            List of dictionaries with parent and student information
        """
        parents = []
        
        if not self.worksheet:
            return parents
            
        # Start from row 2 (skip header), scan column A for parent names
        for row in range(2, self.worksheet.max_row + 1):
            parent_cell = self.worksheet.cell(row=row, column=self.parent_column)
            
            parent_name = str(parent_cell.value).strip() if parent_cell.value else ""
            
            if parent_name:  # Only include rows with parent names
                # Try to get student name from column B (if exists)
                student_name = ""
                if self.worksheet.max_column >= 2:
                    student_cell = self.worksheet.cell(row=row, column=2)
                    student_name = str(student_cell.value).strip() if student_cell.value else ""
                
                parents.append({
                    "row": row,
                    "parent_name": parent_name,
                    "student_name": student_name
                })
        
        return parents
    
    def analyze_month_payments(self, month_name: str, 
                             include_zero_amounts: bool = False,
                             empty_cells_unpaid: bool = True) -> Dict[str, Any]:
        """
        Analyze payments for a specific month
        FIXED: Handle None values properly to avoid comparison errors
        
        Args:
            month_name: Name of month to analyze (e.g., "JUNE", "DECEMBER")
            include_zero_amounts: Whether to treat zero amounts as payments
            empty_cells_unpaid: Whether to treat empty cells as unpaid
            
        Returns:
            Dictionary with analysis results
        """
        # Convert display name to internal name if needed
        if month_name in self.MONTH_DISPLAY.values():
            # Convert "Jun" -> "JUNE"
            month_name = self.MONTH_MAPPING.get(month_name, month_name)
        
        if month_name not in self.column_mapping:
            return {
                "error": f"Month '{month_name}' not found in fee record",
                "paid_parents": [],
                "unpaid_parents": [],
                "total_parents": 0,
                "available_months": self.get_available_months()
            }
        
        month_info = self.column_mapping[month_name]
        date_col = month_info["date_col"]
        amount_col = month_info["amount_col"]
        
        paid_parents = []
        unpaid_parents = []
        
        all_parents = self.get_all_parents()
        
        for parent_info in all_parents:
            row = parent_info["row"]
            parent_name = parent_info["parent_name"]
            student_name = parent_info["student_name"]
            
            # Check date cell
            date_cell = self.worksheet.cell(row=row, column=date_col)
            date_value = str(date_cell.value).strip() if date_cell.value else ""
            
            # Check amount cell
            amount_cell = self.worksheet.cell(row=row, column=amount_col)
            amount_value = self._parse_amount(amount_cell.value)
            
            # Determine payment status based on criteria
            has_date = bool(date_value and date_value.lower() not in ["none", "null", ""])
            
            # FIXED: Handle None values properly
            if amount_value is None:
                has_amount = False
            elif include_zero_amounts:
                has_amount = amount_value >= 0
            else:
                has_amount = amount_value > 0
            
            # If empty cells should be treated as unpaid
            is_empty = (not date_value or date_value.lower() in ["none", "null"]) and amount_value is None
            
            payment_status = {
                "parent_name": parent_name,
                "student_name": student_name,
                "row": row,
                "date_value": date_value,
                "amount_value": amount_value,
                "formatted_amount": self._format_amount(amount_value),
                "has_payment": (has_date or has_amount) and (not empty_cells_unpaid or not is_empty)
            }
            
            if payment_status["has_payment"]:
                paid_parents.append(payment_status)
            else:
                unpaid_parents.append(payment_status)
        
        return {
            "month": month_name,
            "month_display": self.MONTH_DISPLAY.get(month_name, month_name),
            "paid_parents": paid_parents,
            "unpaid_parents": unpaid_parents,
            "total_parents": len(all_parents),
            "paid_count": len(paid_parents),
            "unpaid_count": len(unpaid_parents),
            "column_info": month_info
        }
    
    def analyze_multiple_months(self, month_names: List[str], **kwargs) -> Dict[str, Dict[str, Any]]:
        """
        Analyze payments for multiple months
        
        Args:
            month_names: List of month names to analyze
            **kwargs: Arguments passed to analyze_month_payments
            
        Returns:
            Dictionary with results for each month
        """
        results = {}
        
        for month_name in month_names:
            results[month_name] = self.analyze_month_payments(month_name, **kwargs)
        
        return results
    
    def get_summary_statistics(self) -> Dict[str, Any]:
        """
        Get summary statistics for the entire fee record
        
        Returns:
            Dictionary with overall statistics
        """
        all_parents = self.get_all_parents()
        available_months = self.get_available_months()
        
        return {
            "total_parents": len(all_parents),
            "total_months": len(available_months),
            "available_months": available_months,
            "file_path": self.fee_record_path,
            "last_analyzed": datetime.now().isoformat()
        }
    
    def _parse_amount(self, cell_value) -> Optional[float]:
        """
        Parse amount from cell value, handling various formats
        
        Args:
            cell_value: Raw cell value
            
        Returns:
            Parsed amount as float, or None if not parseable
        """
        if cell_value is None:
            return None
            
        if isinstance(cell_value, (int, float)):
            return float(cell_value)
        
        # Handle string values
        try:
            # Clean common currency symbols and separators
            cleaned = str(cell_value).strip().replace(',', '').replace('$', '').replace('RM', '')
            
            if not cleaned or cleaned.lower() in ['none', 'null', '']:
                return None
                
            return float(cleaned)
            
        except (ValueError, TypeError):
            return None
    
    def _format_amount(self, amount: Optional[float]) -> str:
        """
        Format amount for display
        
        Args:
            amount: Amount to format
            
        Returns:
            Formatted amount string
        """
        if amount is None:
            return ""
        
        if amount == int(amount):
            return str(int(amount))
        else:
            return f"{amount:.2f}".rstrip('0').rstrip('.')
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
            self.workbook = None
            self.worksheet = None